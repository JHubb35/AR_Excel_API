from flask import Flask, send_file
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
from io import BytesIO
import os

app = Flask(__name__)

# API endpoint with preshared token
API_URL = "https://reasolllc.cetecerp.com/api/invoice?invoicedate_from=2023:01:01&preshared_token=8rtpv5gm-dywJH%7C_%5B"

TEMPLATE_FILENAME = "AR_Template_API.xlsx"
CIPS_FILENAME = "cips_data.xlsx"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, TEMPLATE_FILENAME)
CIPS_PATH = os.path.join(BASE_DIR, CIPS_FILENAME)

# ========== Routes ==========
@app.route("/")
def home():
    return "AR Sync API is running. Visit /download to get the Excel report."

@app.route("/download-ar")
def download_excel():
    # --- Step 1: Load API Data ---
    response = requests.get(API_URL)
    response.raise_for_status()
    api_data = response.json()

    # --- Step 2: Load CIPS Data ---
    cips_df = pd.read_excel(CIPS_PATH)
    cips_df["invoice__bc"] = (
        cips_df["invoice__bc"]
        .astype(str)
        .str.strip()
        .str.lower()
    )
    cips_map = {
        row["invoice__bc"]: {
            "receipt_date": row["Invoice Receipt Date"],
            "expected_date": row["Expected Payment Date"]
        }
        for _, row in cips_df.iterrows()
    }

    # --- Step 3: Load Excel Template ---
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # --- Step 4: Write Data to Excel ---
    row_num = 2
    for item in api_data:
        if item.get("ar_status", "").strip().lower() != "open":
            continue

        invoice_bc = str(item.get("invoice__bc", "")).strip().lower()

        ws[f"A{row_num}"] = str(item.get("invoicenum", ""))
        ws[f"B{row_num}"] = item.get("custponum")
        ws[f"C{row_num}"] = item.get("name")
        ws[f"D{row_num}"] = item.get("custnum")
        ws[f"E{row_num}"] = item.get("invoice__bc")
        ws[f"F{row_num}"] = item.get("invoicedate")
        ws[f"G{row_num}"] = item.get("ar_duedate")
        ws[f"H{row_num}"] = item.get("paid_on")
        ws[f"K{row_num}"] = item.get("terms_desc")
        ws[f"S{row_num}"] = item.get("ar_status")

        # Column K - Tax Subtotal
        k_cell = ws[f"M{row_num}"]
        try:
            k_cell.value = float(item.get("total_invoice_amount_in_home_curr", 0))
        except:
            k_cell.value = None
        k_cell.number_format = '"$"#,##0.00'

        
        # Column M - Paid
        m_cell = ws[f"O{row_num}"]
        try:
            m_cell.value = float(item.get("paid", 0))
        except:
            m_cell.value = None
        m_cell.number_format = '"$"#,##0.00'

        # Column O - Total Invoice Amount
        o_cell = ws[f"Q{row_num}"]
        try:
            o_cell.value = float(item.get("total_invoice_amount", 0))
        except:
            o_cell.value = None
        o_cell.number_format = '"$"#,##0.00'


        # CIPS extra fields
        if invoice_bc in cips_map:
            cips_data = cips_map[invoice_bc]
            ws[f"T{row_num}"] = cips_data["receipt_date"]
            ws[f"V{row_num}"] = cips_data["expected_date"]
        else:
            print(f"No match for invoice__bc: {invoice_bc}")    

        # --- Step 5: Formulas ---
        ws.cell(row=row_num, column=9).value  = f'=IF(ISBLANK(F{row_num}), "", TODAY()-F{row_num})'   # I
        ws.cell(row=row_num, column=10).value  = f'=IF(ISBLANK(F{row_num}), "", IF(I{row_num}<30, "<30", IF(I{row_num}<60, ">30", IF(I{row_num}<90, ">60", IF(I{row_num}<120, ">90", IF(I{row_num}<180, ">120", ">180"))))))'  # J
        ws.cell(row=row_num, column=12).value = f'=IF(ISBLANK(G{row_num}), "", TODAY()-G{row_num})'  # L
        ws.cell(row=row_num, column=14).value = f'=IF(ISBLANK(M{row_num}), "", "USD")'               # N
        ws.cell(row=row_num, column=16).value = f'=IF(ISBLANK(O{row_num}), "", "USD")'               # P
        ws.cell(row=row_num, column=18).value = f'=IF(ISBLANK(Q{row_num}), "", "USD")'               # R
        ws.cell(row=row_num, column=21).value = f'=IF(A{row_num}="", "", IF(T{row_num}="", "No", "Yes"))'  # U

        row_num += 1

    # --- Step 6: Return as Download ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="New_synced_output.xlsx", as_attachment=True)

# ========== Entry Point ==========
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))  # Railway will inject PORT
    app.run(host="0.0.0.0", port=port)
